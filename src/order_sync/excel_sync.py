import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.workbook.protection import WorkbookProtection
from datetime import datetime


COLUMNS = [
	"order_id",
	"date",
	"status",
	"total",
	"updated_at",
	"data_json",
]
SHEET_NAME = "orders"


# Report sheet for Mongo mapping
REPORT_SHEET_NAME = "report"
META_SHEET_NAME = "meta"


def ensure_workbook(path: Path) -> Workbook:
	if path.exists():
		wb = load_workbook(filename=str(path))
		# keep meta hidden if present
		if META_SHEET_NAME in wb.sheetnames:
			wb[META_SHEET_NAME].sheet_state = "veryHidden"
			wb.security = WorkbookProtection(lockStructure=True)
		return wb
	# Create a new workbook with the default sheet (do not rename or add 'orders' headers here)
	wb = Workbook()
	# lock structure by default
	wb.security = WorkbookProtection(lockStructure=True)
	return wb


def get_orders_sheet(wb: Workbook) -> Worksheet:
	if SHEET_NAME in wb.sheetnames:
		return wb[SHEET_NAME]
	ws = wb.create_sheet(SHEET_NAME)
	# add headers if sheet is new/empty
	if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
		ws.append(COLUMNS)
	return ws


def sheet_headers_index(sheet: Worksheet) -> Dict[str, int]:
	headers: Dict[str, int] = {}
	for idx, cell in enumerate(sheet[1], start=1):
		if cell.value:
			headers[str(cell.value)] = idx
	return headers


def find_order_row(sheet: Worksheet, order_id: str, headers: Dict[str, int]) -> Optional[int]:
	col_idx = headers.get("order_id")
	if not col_idx:
		return None
	for row in range(2, sheet.max_row + 1):
		val = sheet.cell(row=row, column=col_idx).value
		if val == order_id:
			return row
	return None


def write_order_row(sheet: Worksheet, row: int, data: Dict[str, Any], headers: Dict[str, int]) -> None:
	for key in ["order_id", "date", "status", "total", "updated_at", "data_json"]:
		col = headers.get(key)
		if not col:
			continue
		value = data.get(key)
		sheet.cell(row=row, column=col, value=value)


def upsert_orders_for_user(user_id: str, orders: List[Dict[str, Any]], output_dir: Path) -> Path:
	output_dir.mkdir(parents=True, exist_ok=True)
	wb_path = output_dir / f"{user_id}.xlsx"
	wb = ensure_workbook(wb_path)
	sheet = get_orders_sheet(wb)
	headers = sheet_headers_index(sheet)
	# ensure all columns exist
	if not all(h in headers for h in COLUMNS):
		missing = [h for h in COLUMNS if h not in headers]
		for h in missing:
			sheet.cell(row=1, column=sheet.max_column + 1, value=h)
		headers = sheet_headers_index(sheet)

	for o in orders:
		order_id = str(o["orderId"]).strip()
		data_row = {
			"order_id": order_id,
			"date": o.get("date"),
			"status": o.get("status"),
			"total": o.get("total"),
			"updated_at": o.get("updatedAt"),
			"data_json": json.dumps({k: v for k, v in o.items() if k not in {"orderId", "userId"}}, ensure_ascii=False),
		}
		existing_row = find_order_row(sheet, order_id, headers)
		if existing_row is None:
			# append at the end
			row_index = sheet.max_row + 1
			write_order_row(sheet, row_index, data_row, headers)
		else:
			# update
			write_order_row(sheet, existing_row, data_row, headers)

	wb.save(str(wb_path))
	return wb_path


def read_last_sync(path: Path) -> Optional[datetime]:
	if not path.exists():
		return None
	wb = load_workbook(filename=str(path))
	if META_SHEET_NAME not in wb.sheetnames:
		return None
	ws = wb[META_SHEET_NAME]
	# expect cell A1 header, A2 value ISO
	val = ws.cell(row=2, column=1).value
	try:
		return datetime.fromisoformat(val) if val else None
	except Exception:
		return None


def write_last_sync(path: Path, when: datetime) -> None:
	wb = ensure_workbook(path)
	if META_SHEET_NAME in wb.sheetnames:
		ws = wb[META_SHEET_NAME]
		for row in ws[2:ws.max_row]:
			pass
	else:
		ws = wb.create_sheet(META_SHEET_NAME)
		ws.append(["last_sync"])
	# write/update cell A2
	ws.cell(row=2, column=1, value=when.isoformat())
	# hide meta and lock workbook structure
	ws.sheet_state = "veryHidden"
	wb.security = WorkbookProtection(lockStructure=True)
	wb.save(str(path))


def _auto_size_columns(ws: Worksheet) -> None:
	for col in ws.columns:
		max_length = 0
		col_letter = col[0].column_letter
		for cell in col:
			val = cell.value
			length = len(str(val)) if val is not None else 0
			if length > max_length:
				max_length = length
		ws.column_dimensions[col_letter].width = min(max(10, max_length + 2), 60)


def _apply_date_formats(ws: Worksheet, date_headers: List[str]) -> None:
	# Apply a simple date format to columns matching given headers
	headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
	for h in date_headers:
		idx = headers.get(h)
		if not idx:
			continue
		for r in range(2, ws.max_row + 1):
			cell = ws.cell(row=r, column=idx)
			if isinstance(cell.value, str) and len(cell.value) in (10, 19):
				# leave as text; Excel will still display it; robust parsing is environment-specific
				cell.alignment = Alignment(horizontal="left")


def _hide_columns(ws: Worksheet, headers_to_hide: List[str]) -> None:
	headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
	for h in headers_to_hide:
		idx = headers.get(h)
		if idx:
			ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].hidden = True


def _format_report_sheet(ws: Worksheet, columns: List[str]) -> None:
	# Freeze header
	ws.freeze_panes = "A2"
	# Create Excel table over the used range
	end_col_letter = ws.cell(row=1, column=len(columns)).column_letter
	end_row = ws.max_row
	table = Table(displayName="ReportTable", ref=f"A1:{end_col_letter}{end_row}")
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
	table.tableStyleInfo = style
	ws.add_table(table)
	# Auto-size columns and basic date formatting
	_auto_size_columns(ws)
	_apply_date_formats(ws, [
		"ETD (fecha)",
		"ETA (fecha)",
		"Fecha de ISF",
	])
	# Hide technical columns
	_hide_columns(ws, ["orderId", "__createdAt", "__lastUpdateAt"])


def write_report_for_user(user_id: str, rows: List[Dict[str, Any]], columns: List[str], output_dir: Path) -> Path:
	"""Create or overwrite a 'report' sheet with the provided rows and column order. Remove other sheets.
	Adds table styling, frozen header, basic date formatting, and auto-size columns."""
	output_dir.mkdir(parents=True, exist_ok=True)
	wb_path = output_dir / f"{user_id}.xlsx"
	wb = ensure_workbook(wb_path)
	# Drop existing report sheet (if any)
	if REPORT_SHEET_NAME in wb.sheetnames:
		ws_old = wb[REPORT_SHEET_NAME]
		wb.remove(ws_old)
	# Create report sheet and write rows
	ws = wb.create_sheet(REPORT_SHEET_NAME)
	ws.append(columns)
	for r in rows:
		ws.append([r.get(c) for c in columns])
	_format_report_sheet(ws, columns)
	# Remove any other sheet to keep only 'report'
	for name in list(wb.sheetnames):
		if name != REPORT_SHEET_NAME and name != META_SHEET_NAME:
			wb.remove(wb[name])
	wb.save(str(wb_path))
	return wb_path


def read_report_rows(path: Path, columns: List[str]) -> List[Dict[str, Any]]:
	"""Read current report rows into a list of dicts (without headers)."""
	if not path.exists():
		return []
	wb = load_workbook(filename=str(path))
	if REPORT_SHEET_NAME not in wb.sheetnames:
		return []
	ws = wb[REPORT_SHEET_NAME]
	rows: List[Dict[str, Any]] = []
	for r in range(2, ws.max_row + 1):
		row: Dict[str, Any] = {}
		for c_idx, col_name in enumerate(columns, start=1):
			row[col_name] = ws.cell(row=r, column=c_idx).value
		rows.append(row)
	return rows


def write_report_rows(path: Path, columns: List[str], rows: List[Dict[str, Any]]) -> Path:
	"""Overwrite report sheet with given rows and reapply formatting."""
	wb = ensure_workbook(path)
	if REPORT_SHEET_NAME in wb.sheetnames:
		wb.remove(wb[REPORT_SHEET_NAME])
	ws = wb.create_sheet(REPORT_SHEET_NAME)
	ws.append(columns)
	for r in rows:
		ws.append([r.get(c) for c in columns])
	_format_report_sheet(ws, columns)
	# Remove any other non-meta sheet
	for name in list(wb.sheetnames):
		if name != REPORT_SHEET_NAME and name != META_SHEET_NAME:
			wb.remove(wb[name])
	wb.save(str(path))
	return path


def upsert_report_for_user(user_id: str, changed_rows: List[Dict[str, Any]], columns: List[str], output_dir: Path) -> Path:
	"""Merge changed rows into existing report by orderId and write back, preserving order by __createdAt asc."""
	output_dir.mkdir(parents=True, exist_ok=True)
	wb_path = output_dir / f"{user_id}.xlsx"
	existing = read_report_rows(wb_path, columns)
	by_id: Dict[str, Dict[str, Any]] = {}
	for r in existing:
		key = str(r.get("orderId")) if r.get("orderId") is not None else None
		if key:
			by_id[key] = r
	# Apply changes
	for r in changed_rows:
		key = str(r.get("orderId")) if r.get("orderId") is not None else None
		if not key:
			continue
		by_id[key] = r
	# Build merged and sort by __createdAt asc (missing last)
	merged = list(by_id.values())
	merged.sort(key=lambda x: (x.get("__createdAt") or ""))
	return write_report_rows(wb_path, columns, merged)


def upsert_report_for_user_with_stats(user_id: str, changed_rows: List[Dict[str, Any]], columns: List[str], output_dir: Path) -> Tuple[Path, List[str], List[str]]:
	"""Like upsert_report_for_user, but returns (wb_path, created_ids, updated_ids)."""
	output_dir.mkdir(parents=True, exist_ok=True)
	wb_path = output_dir / f"{user_id}.xlsx"
	existing = read_report_rows(wb_path, columns)
	existing_ids = set()
	by_id: Dict[str, Dict[str, Any]] = {}
	for r in existing:
		key = str(r.get("orderId")) if r.get("orderId") is not None else None
		if key:
			existing_ids.add(key)
			by_id[key] = r
	created_ids: List[str] = []
	updated_ids: List[str] = []
	for r in changed_rows:
		key = str(r.get("orderId")) if r.get("orderId") is not None else None
		if not key:
			continue
		if key in existing_ids:
			updated_ids.append(key)
		else:
			created_ids.append(key)
		by_id[key] = r
	merged = list(by_id.values())
	merged.sort(key=lambda x: (x.get("__createdAt") or ""))
	path = write_report_rows(wb_path, columns, merged)
	return path, created_ids, updated_ids
