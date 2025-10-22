import argparse
import sys
from pathlib import Path
from typing import Dict, List
from datetime import datetime, timezone
import tempfile

from openpyxl import load_workbook

from .excel_sync import write_report_for_user, read_last_sync, write_last_sync, upsert_report_for_user_with_stats
from .config import load_env_file, get_config
from .drive import build_drive_client, upload_or_update_file, find_file_id_by_name, download_file
from .mongo_fetch import (
	fetch_orders_by_account,
	fetch_account_name,
	fetch_updated_order_ids_since,
	fetch_orders_by_ids,
	fetch_updated_logs_since,
	fetch_recent_field_changes,
)
from .mongo_mapping import map_doc_to_report_row, REPORT_COLUMNS


def _append_log(output_dir: Path, line: str) -> None:
	try:
		output_dir.mkdir(parents=True, exist_ok=True)
		log_path = output_dir / "order_sync.log"
		with open(log_path, "a", encoding="utf-8") as f:
			f.write(line + "\n")
	except Exception:
		pass


def cmd_mongo_auto(args: argparse.Namespace) -> int:
	cfg = get_config()
	if not cfg.mongo_uri:
		print("ERROR: MONGO_URI must be set in env/.env", file=sys.stderr)
		return 2

	drive_client = None
	if cfg.drive_client_email and cfg.drive_private_key and cfg.drive_folder_id:
		try:
			drive_client = build_drive_client(cfg.drive_client_email, cfg.drive_private_key)
		except Exception as e:
			print(f"WARN: Could not initialize Drive client: {e}", file=sys.stderr)
			drive_client = None

	if not drive_client or not cfg.drive_folder_id:
		print("ERROR: Drive must be configured (GOOGLE_CLIENT_EMAIL, GOOGLE_PRIVATE_KEY, GOOGLE_DRIVE_FOLDER_ID)", file=sys.stderr)
		return 2

	account_ids: List[str] = []
	if args.account_id:
		account_ids = [args.account_id]
	elif cfg.account_ids:
		account_ids = cfg.account_ids
	else:
		print("ERROR: Provide --account-id or set ACCOUNT_IDS in env", file=sys.stderr)
		return 2

	output_dir = Path(args.output_dir or cfg.output_dir)
	utc_now = datetime.now(timezone.utc).astimezone(timezone.utc).replace(tzinfo=None)
	for acc_id in account_ids:
		acc_name = fetch_account_name(cfg.mongo_uri, acc_id)
		filename_id = acc_name if acc_name else acc_id
		remote_name = f"{filename_id}.xlsx"
		file_id = find_file_id_by_name(drive_client, cfg.drive_folder_id, remote_name)
		with tempfile.TemporaryDirectory() as tmpdir:
			tmp_path = Path(tmpdir) / remote_name
			if file_id:
				# download and decide full vs incremental based on 'report' sheet presence
				try:
					download_file(drive_client, file_id, tmp_path)
				except Exception:
					pass
				has_report = False
				try:
					wb = load_workbook(filename=str(tmp_path))
					has_report = ("report" in wb.sheetnames)
				except Exception:
					has_report = False
				if not has_report:
					# full rebuild if report sheet is missing
					all_docs = fetch_orders_by_account(cfg.mongo_uri, acc_id)
					all_rows = [map_doc_to_report_row(d) for d in all_docs]
					if acc_id not in cfg.account_ids_no_prefix and cfg.ref_prefixes:
						all_rows = [r for r in all_rows if isinstance(r.get("REF"), str) and any(r.get("REF", "").startswith(p) for p in cfg.ref_prefixes)]
					wb_path = write_report_for_user(filename_id, all_rows, REPORT_COLUMNS, Path(tmpdir))
					write_last_sync(wb_path, utc_now)
					msg = f"[{utc_now.isoformat()}] Full generated for {filename_id}: created={len(all_rows)}, updated=0"
					print(msg)
					_append_log(output_dir, msg)
					_, action = upload_or_update_file(drive_client, wb_path, cfg.drive_folder_id)
					print(f"Drive {action}: {remote_name}")
					continue
				# incremental flow
				last_sync = read_last_sync(tmp_path) or (utc_now.replace(year=utc_now.year - 1))
				order_ids = fetch_updated_order_ids_since(cfg.mongo_uri, acc_id, since=last_sync)
				if not order_ids:
					msg = f"[{utc_now.isoformat()}] No updates for {filename_id} since {last_sync.isoformat()}"
					print(msg)
					_append_log(output_dir, msg)
					continue
				docs = fetch_orders_by_ids(cfg.mongo_uri, order_ids)
				changed_rows = [map_doc_to_report_row(d) for d in docs]
				if acc_id not in cfg.account_ids_no_prefix and cfg.ref_prefixes:
					changed_rows = [r for r in changed_rows if isinstance(r.get("REF"), str) and any(r.get("REF", "").startswith(p) for p in cfg.ref_prefixes)]
				wb_path, created_ids, updated_ids = upsert_report_for_user_with_stats(filename_id, changed_rows, REPORT_COLUMNS, Path(tmpdir))
				write_last_sync(wb_path, utc_now)
				msg = f"[{utc_now.isoformat()}] Incremental for {filename_id}: created={len(created_ids)}, updated={len(updated_ids)}"
				print(msg)
				_append_log(output_dir, msg)
				if updated_ids:
					changes = fetch_recent_field_changes(cfg.mongo_uri, acc_id, since=last_sync, order_ids=updated_ids[:50])
					for oid, entries in changes:
						for e in entries:
							_append_log(output_dir, f"  ~ {oid} {e.get('action')} @ {e.get('date')}: " + ", ".join([f"{c.get('field')}: {c.get('old')} -> {c.get('new')}" for c in e.get('changes', [])]))
				_, action = upload_or_update_file(drive_client, wb_path, cfg.drive_folder_id)
				print(f"Drive {action}: {remote_name}")
			else:
				# no file in Drive → full
				all_docs = fetch_orders_by_account(cfg.mongo_uri, acc_id)
				all_rows = [map_doc_to_report_row(d) for d in all_docs]
				if acc_id not in cfg.account_ids_no_prefix and cfg.ref_prefixes:
					all_rows = [r for r in all_rows if isinstance(r.get("REF"), str) and any(r.get("REF", "").startswith(p) for p in cfg.ref_prefixes)]
				wb_path = write_report_for_user(filename_id, all_rows, REPORT_COLUMNS, Path(tmpdir))
				write_last_sync(wb_path, utc_now)
				msg = f"[{utc_now.isoformat()}] Full generated for {filename_id}: created={len(all_rows)}, updated=0"
				print(msg)
				_append_log(output_dir, msg)
				_, action = upload_or_update_file(drive_client, wb_path, cfg.drive_folder_id)
				print(f"Drive {action}: {remote_name}")

	print(f"Auto sync processed {len(account_ids)} account(s)")
	return 0


def build_parser() -> argparse.ArgumentParser:
	p = argparse.ArgumentParser(prog="order-sync", description="Sync orders into per-user Excel workbooks (Mongo auto), Google Drive as source of truth")
	p.add_argument("--env-file", help="Path to .env file to load", default=None)
	sub = p.add_subparsers(dest="cmd", required=True)

	pa = sub.add_parser("mongo-auto", help="Full on first run (if not in Drive), incremental otherwise")
	pa.add_argument("--account-id", help="Mongo ObjectId of account. If omitted, reads ACCOUNT_IDS from env.")
	pa.add_argument("--output-dir", default=None)
	pa.add_argument("--verbose", action="store_true")
	pa.set_defaults(func=cmd_mongo_auto)
	return p


def main() -> None:
	parser = build_parser()
	args = parser.parse_args()
	load_env_file(args.env_file)
	code = args.func(args)
	sys.exit(code)
